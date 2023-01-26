DATA_P = [
        [{'sede': 'Avenida 68', 'ciclo': 'Ciclo II Fin de Semana Motricidad 2023 Calendario A', 'nivel': 'Salsa básico', 'de': '06:30:00', 'a': '07:30:00', 'dias': 'Sábados y Domingos'},
        {'sede': 'Avenida 68', 'ciclo': 'Ciclo II Fin de Semana Motricidad 2023 Calendario A', 'nivel': 'Baile básico', 'de': '07:00:00', 'a': '08:00:00', 'dias': 'Sábados'},
        {'sede': 'Avenida 68', 'ciclo': 'Ciclo II Fin de Semana Motricidad 2023 Calendario A', 'nivel': 'Salsa básico', 'de': '07:00:00', 'a': '08:00:00', 'dias': 'Sábados'}
        ],
        [{'sede': 'Cajicá', 'ciclo': 'Ciclo II Fin de Semana Motricidad 2023 Calendario A', 'nivel': 'Rumba kids - 7 a 12 años', 'de': '08:00:00', 'a': '09:00:00', 'dias': 'Sábados'},
        {'sede': 'Cajicá', 'ciclo': 'Ciclo II Fin de Semana Motricidad 2023 Calendario A', 'nivel': 'Ballet básico - 7 a 12 años', 'de': '09:00:00', 'a': '10:00:00', 'dias': 'Sábados'},
        {'sede': 'Cajicá', 'ciclo': 'Ciclo II Fin de Semana Motricidad 2023 Calendario A', 'nivel': 'Rumba kids - 4 a 6 años', 'de': '10:00:00', 'a': '11:00:00', 'dias': 'Sábados'},
        ]
    ],


DATA_S = [
       [{'inscritos': 0, 'SKU': 2520181},
        {'inscritos': 0, 'SKU': 2520117},
        {'inscritos': 0, 'SKU': 2520116},
        ],
       [{'inscritos': 0, 'SKU': 2520291},
        {'inscritos': 0, 'SKU': 2520292},
        {'inscritos': 0, 'SKU': 2520293},
        {'inscritos': 0, 'SKU': 2520294},
        ]
    ]


for l in range(0, len(DATA_P[0])):
    print(DATA_P[0][l])
    print(l)

# excel


import openpyxl
# from index import obj

DATA_P = [
        [{'sede': 'Avenida 68', 'ciclo': 'Ciclo II Fin de Semana Motricidad 2023 Calendario A', 'nivel': 'Salsa básico', 'de': '06:30:00', 'a': '07:30:00', 'dias': 'Sábados y Domingos'},
        {'sede': 'Avenida 68', 'ciclo': 'Ciclo II Fin de Semana Motricidad 2023 Calendario A', 'nivel': 'Baile básico', 'de': '07:00:00', 'a': '08:00:00', 'dias': 'Sábados'},
        {'sede': 'Avenida 68', 'ciclo': 'Ciclo II Fin de Semana Motricidad 2023 Calendario A', 'nivel': 'Salsa básico', 'de': '07:00:00', 'a': '08:00:00', 'dias': 'Sábados'}
        ],
        [{'sede': 'Cajicá', 'ciclo': 'Ciclo II Fin de Semana Motricidad 2023 Calendario A', 'nivel': 'Rumba kids - 7 a 12 años', 'de': '08:00:00', 'a': '09:00:00', 'dias': 'Sábados'},
        {'sede': 'Cajicá', 'ciclo': 'Ciclo II Fin de Semana Motricidad 2023 Calendario A', 'nivel': 'Ballet básico - 7 a 12 años', 'de': '09:00:00', 'a': '10:00:00', 'dias': 'Sábados'},
        {'sede': 'Cajicá', 'ciclo': 'Ciclo II Fin de Semana Motricidad 2023 Calendario A', 'nivel': 'Rumba kids - 4 a 6 años', 'de': '10:00:00', 'a': '11:00:00', 'dias': 'Sábados'},
        ]
    ],


DATA_S = [
       [{'inscritos': 0, 'SKU': 2520181},
        {'inscritos': 0, 'SKU': 2520117},
        {'inscritos': 0, 'SKU': 2520116},
        ],
       [{'inscritos': 0, 'SKU': 2520291},
        {'inscritos': 0, 'SKU': 2520292},
        {'inscritos': 0, 'SKU': 2520293},
        {'inscritos': 0, 'SKU': 2520294},
        ]
    ]
# Iniciado sesión y estrayendo datos
# login = obj.login()
# all_data = obj.loop_api()
# DATA_P = obj.data_p
# DATA_S = obj.data_s

# Cargando archiivo de excel
workbook=openpyxl.load_workbook("app/programacion.xlsx")
shenames=workbook.get_sheet_names()
worksheet=workbook[shenames[2]]
name = worksheet.title
rows=worksheet.max_row
columns=worksheet.max_column

# Variables necesarias
data_header = [
    'operación',
    'deporte',
    'sede',
    'categoría',
    'ciclo',
    'nivel',
    'escenario',
    'zona',
    'de',
    'a',
    'dias',
    'fecha inicial',
    'fecha final',
    'rama',
    'cupo',
    'inscritos',
    'SKU',
    'observaciones'
    ]

datos_omitidos = [
    "operación",
    "zona",
    'fecha inicial',
    'fecha final',
    'cupo',
    'observaciones',
    'escenario',
    'categoría',
    'deporte',
    'rama',
    "inscritos",
    "SKU"
    ]

data_excel = []
data_error = []

for i in range(1,rows):
    data_excel.append({})

# Estrayendo data de excel
for row, i in zip(worksheet.rows, range(0,rows-1)):
    for cell, p in zip(row, data_header):
        if p in datos_omitidos:
            continue
        else:
            data_excel[i][p] = str(cell.value)

# Insertando tados en excel
def insertar_datos(excel, dato, posicion):
    excel[f"P{posicion}"] = dato["inscritos"]
    excel[f"Q{posicion}"] = dato["SKU"]


def comparar(data1_, excel_, inidice_DATAP_M):
    for indice, valor in zip(range(1, len(excel_)+2), excel_):
        if data1_["valor"] == valor:
            # Insertando valor a excel
            insertar_datos(excel=worksheet, dato = DATA_S[inidice_DATAP_M][data1_['indice']], posicion = indice)
            # print(f"Los datos coinciden, el codido de data_hercules es: {data1['indice']} - El valor del codigo de excel es: {indice}" )
            return True

# for indice, valor in zip(range(0, len(data_1)), data_1):
#     if comparar(data1 = {"indice": indice, "valor": valor}, excel = data_excel) != True:
#         # print("El dato no se encontró en excel: ", data_1[indice])
#         data_error.append(data_1[indice])

def interacion_mmayor(data_1_, data_excel_, data_error_, inidice_DATAP_):
    for indice, valor in zip(range(0, len(data_1_)), data_1_):
        if comparar(data1_ = {"indice": indice, "valor": valor}, excel = data_excel_, inidice_DATAP_M = inidice_DATAP_) != True:
            # print("El dato no se encontró en excel: ", data_1[indice])
            data_error_.append(data_1_[indice])

# for inidice_DATAP, l in zip(DATA_P, range(0, len(DATA_P))):
#     interacion_mmayor(DATA_P[inidice_DATAP], data_excel_= data_excel, data_error_=data_error, inidice_DATAP_ = l)

for inidice_DATAP, l in zip(DATA_P, range(0, len(DATA_P))):
    print(inidice_DATAP)
    print(l)

# Guardando el archivo de excel
workbook.save(filename = "app/programacion.xlsx")

# Imprimiendo datos con errores
print(data_error)


# index

import requests
from bs4 import BeautifulSoup
import urllib3
urllib3.disable_warnings()


class SesionActive():
    data_p = []
    data_s = []
    sedes_input_api = {
        "fin_semana_calenadarioA_av68": {"idDeporte": "544","idCiclo": "26554","idSede": "476","start": "0","limit": "100"},
        "fin_semana_calenadarioA_cajica": {"idDeporte": "544", "idCiclo": "26554", "idSede": "4", "start": "0", "limit": "100"},
        "fin_semana_calenadarioA_220": {'idDeporte': '544', 'idCiclo': '26554', 'idSede': '2', 'start': '0', 'limit': '100'},
        "fin_semana_calenadarioA_cll94": {'idDeporte': '544', 'idCiclo': '26554', 'idSede': '5', 'start': '0', 'limit': '100'},
        "fin_semana_calenadarioA_cbi_americas": {'idDeporte': '544', 'idCiclo': '26554', 'idSede': '588', 'start': '0', 'limit': '100'},
        "fin_semana_calenadarioA_cbi_autosur": {'idDeporte': '544', 'idCiclo': '26554', 'idSede': '606', 'start': '0', 'limit': '100'},
        "fin_semana_calenadarioA_cbi_142": {'idDeporte': '544', 'idCiclo': '26554', 'idSede': '610', 'start': '0', 'limit': '100'},
        "fin_semana_calenadarioA_cbi_centro_mayor": {'idDeporte': '544', 'idCiclo': '26554', 'idSede': '689', 'start': '0', 'limit': '100'},
        "fin_semana_calenadarioA_cbi_soacha": {'idDeporte': '544', 'idCiclo': '26554', 'idSede': '607', 'start': '0', 'limit': '100'},
        "fin_semana_calenadarioA_cbi_zona_franca": {'idDeporte': '544', 'idCiclo': '26554', 'idSede': '609', 'start': '0', 'limit': '100'},
        "fin_semana_calenadarioA_cef": {'idDeporte': '544', 'idCiclo': '26554', 'idSede': '573', 'start': '0', 'limit': '100'},
        "fin_semana_calenadarioA_cur": {'idDeporte': '544', 'idCiclo': '26554', 'idSede': '3', 'start': '0', 'limit': '100'},
        "fin_semana_calenadarioA_online": {'idDeporte': '544', 'idCiclo': '26554', 'idSede': '584', 'start': '0', 'limit': '100'},
        "fin_semana_calenadarioA_suba": {'idDeporte': '544', 'idCiclo': '26554', 'idSede': '584', 'start': '0', 'limit': '100'},
        "fin_semana_calenadarioB_av68": {"idDeporte": "544","idCiclo": "26555","idSede": "476","start": "0","limit": "100"},
        "fin_semana_calenadarioB_cbi_soacha": {'idDeporte': '544', 'idCiclo': '26555', 'idSede': '607', 'start': '0', 'limit': '100'},
        "entre_semana_calenadarioB_av68": {"idDeporte": "544","idCiclo": "26556","idSede": "476","start": "0","limit": "100"},
        "entre_semana_calenadarioB_call94": {"idDeporte": "5","idCiclo": "26556","idSede": "476","start": "0","limit": "100"},
        "entre_semana_calenadarioB_cbi_americas": {"idDeporte": "588","idCiclo": "26556","idSede": "476","start": "0","limit": "100"},
        "entre_semana_calenadarioB_cbi_cll142": {"idDeporte": "610","idCiclo": "26556","idSede": "476","start": "0","limit": "100"},
        "entre_semana_calenadarioB_cbi_centro_mayor": {"idDeporte": "689","idCiclo": "26556","idSede": "476","start": "0","limit": "100"},
        "entre_semana_calenadarioB_cbi_pasus_veraguas": {"idDeporte": "618","idCiclo": "26556","idSede": "476","start": "0","limit": "100"},
        "entre_semana_calenadarioB_cbi_soacha": {"idDeporte": "607","idCiclo": "26556","idSede": "476","start": "0","limit": "100"},
        "entre_semana_calenadarioB_cur": {"idDeporte": "3","idCiclo": "26556","idSede": "476","start": "0","limit": "100"},
        "entre_semana_calenadarioB_online": {"idDeporte": "620","idCiclo": "26556","idSede": "476","start": "0","limit": "100"},
        "entre_semana_calenadarioB_suba": {"idDeporte": "584","idCiclo": "26556","idSede": "476","start": "0","limit": "100"},
        }
    SESSION = requests.Session()
    SESSION.verify = False
    data_login = {
        "usuario": "jmfialloch",
        "password": "Juan060997."
        }
    dias = {
        "lun": "Lunes",
        "mar": "Martes",
        "mié": "Miércoles",
        "jue": "Jueves",
        "vie": "Viernes",
        "sáb": "Sábados",
        "dom": "Domingos",
        "sáb y dom": "Sábados y Domingos"
        }
    headers = {
        'authority': 'sistemamotricidad.deportescompensar.com',
        'accept': 'application/json, text/plain, */*',
        'accept-language': 'es-419,es;q=0.9,es-ES;q=0.8,en;q=0.7,en-GB;q=0.6,en-US;q=0.5',
        'cache-control': 'no-cache',
        'content-type': 'application/json;charset=UTF-8',
        # 'cookie': '_ga=GA1.2.1015349931.1649043978; _gid=GA1.2.143685208.1674574066; sistema=1ttpt51iai985lng117f5eju5alk8uur',
        'origin': 'https://sistemamotricidad.deportescompensar.com',
        'pragma': 'no-cache',
        'referer': 'https://sistemamotricidad.deportescompensar.com/sistema.php/login',
        'sec-ch-ua': '"Not_A Brand";v="99", "Microsoft Edge";v="109", "Chromium";v="109"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"macOS"',
        'sec-fetch-dest': 'empty',
        'sec-fetch-mode': 'cors',
        'sec-fetch-site': 'same-origin',
        'user-agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.0.0 Safari/537.36 Edg/109.0.1518.61',
        }

    def login(self):
        url_login = "https://sistemamotricidad.deportescompensar.com/sistema.php/login"
        login = self.SESSION.post(url = url_login,  json = {"usuario": "jmfialloch", "password": "Juan060997."}, headers=self.headers, verify=False)
        soup = BeautifulSoup(login.content, "html.parser")
        print(soup)

    def extract_data_api(self, code_):
        data_principal = []
        data_secundaria = []
        url = "https://sistemamotricidad.deportescompensar.com/sistema.php/v1/pruebacampeonato/json"
        code = code_
        # code = {"idDeporte": "544","idCiclo": "26554","idSede": "476","start": "0","limit": "100"}
        respuesta = self.SESSION.post(url, code, verify=False)
        dia = ""
        for i in respuesta.json()["pruebas"]:
            data_principal.append({
                "sede":i["sede"],
                "ciclo": i["ciclo"],
                "nivel":i["prueba"].rstrip(i["prueba"][-1]),
                "de":i["horario"][0:5]+":00",
                "a":i["horario"][6:11]+":00",
                "dias":i["dias"]
            })
            dia = self.dias[data_principal[len(data_principal) - 1]['dias']]
            data_principal[len(data_principal) - 1]['dias'] = dia

        for i in respuesta.json()["pruebas"]:
            data_secundaria.append({
                "inscritos":i["num_participantes"],
                "SKU":i["id"],
            })
        self.data_p.append(data_principal)
        self.data_s.append(data_secundaria)
        return [data_principal, data_secundaria]
    
    def loop_api(self):
            for i, c in zip(self.sedes_input_api, range(0,3)):
                if c <= 2:
                    self.extract_data_api(self.sedes_input_api[i])
                    print("s")
                else: 
                    break
        
            # def loop_api(self):
    #     for i in self.sedes_input_api:
    #         self.extract_data_api(self.sedes_input_api[i])
    #         print("s")

obj = SesionActive()
obj.login()
data_1 = obj.loop_api()
print("obj.data_p[0]", obj.data_p[0])
print("obj.data_p[1]", obj.data_p[1])
print("obj.data_s[0]", obj.data_s[0])
print("obj.data_s[1]", obj.data_s[1])