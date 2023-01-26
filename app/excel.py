import openpyxl
from index import obj

# Iniciado sesión y estrayendo datos
login = obj.login()
data_1 = obj.Av68_CicloII_FinSemana_CalendarioA()[0]
data_2 = obj.Av68_CicloII_FinSemana_CalendarioA()[1]

# Cargando archiivo de excel
workbook=openpyxl.load_workbook("app/programacion.xlsx")
shenames=workbook.get_sheet_names()
worksheet=workbook[shenames[2]]
name = worksheet.title
rows=worksheet.max_row
columns=worksheet.max_column

# Variables necesarias
data_excel = []
data_header = [
            'operación',
            'deporte',
            'sede',
            'categoría',
            'ciclo',
            'nivel',
            'escenario',
            'zona',
            'de',
            'a',
            'dias',
            'fecha inicial',
            'fecha final',
            'rama',
            'cupo',
            'inscritos',
            'SKU',
            'observaciones'
            ]
data_error = []
datos_omitidos = [
    "operación", 
    "zona", 
    'fecha inicial', 'fecha final', 'cupo', 'observaciones', 'escenario', 'categoría', 'deporte', 'rama', "inscritos", "SKU"]

for i in range(1,rows):
    data_excel.append({})

# Estrayendo data de excel
for row, i in zip(worksheet.rows, range(0,rows-1)):
    for cell, p in zip(row, data_header):
        if p in ["operación", "zona", 'fecha inicial', 'fecha final', 'cupo', 'observaciones', 'escenario', 'categoría', 'deporte', 'rama', "inscritos", "SKU"]:
            continue
        else:
            data_excel[i][p] = str(cell.value)

# Insertando tados en excel
def insertar_datos(excel, dato, posicion):
    excel[f"P{posicion}"] = dato["inscritos"]
    excel[f"Q{posicion}"] = dato["SKU"]


def comparar(data1, excel):
    for indice, valor in zip(range(1, len(excel)+2), excel):
        if data1["valor"] == valor:
            # Insertando valor a excel
            insertar_datos(excel=worksheet, dato = data_2[data1['indice']], posicion = indice)
            # print(f"Los datos coinciden, el codido de data_hercules es: {data1['indice']} - El valor del codigo de excel es: {indice}" )
            return True

for indice, valor in zip(range(0, len(data_1)), data_1):
    if comparar(data1 = {"indice": indice, "valor": valor}, excel = data_excel) != True:
        # print("El dato no se encontró en excel: ", data_1[indice])
        data_error.append(data_1[indice])

# Guardando el archivo de excel
workbook.save(filename = "app/programacion.xlsx")

# Imprimiendo datos con errores
print(data_error)
